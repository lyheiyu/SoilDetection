# import pandas as pd
#
# # 1. 定义每个 Sheet 对应的列结构
#
# # 指标表（Indicators）
# columns_indicators = [
#     "Indicator ID",
#     "Indicator Name",
#     "Indicator Code",
#     "Unit",
#     "Data Source",
#     "Category",
#     "Availability",
#     "Notes"
# ]
#
# # 国家表（Countries）
# columns_countries = [
#     "Country Code",
#     "Country Name",
#     "Region",
#     "ISO Code",
#     "Notes"
# ]
#
# # 观测值表（Observations）
# columns_observations = [
#     "Indicator ID",
#     "Country Code",
#     "Year",
#     "Value",
#     "Notes"
# ]
#
# # 2. 创建空的 DataFrame
# df_indicators = pd.DataFrame(columns=columns_indicators)
# df_countries = pd.DataFrame(columns=columns_countries)
# df_observations = pd.DataFrame(columns=columns_observations)
#
# # 3. 将三个 DataFrame 写入同一个 Excel 文件，不同 Sheet
# output_file = "Indicators_MultiSheet_Template.xlsx"
#
# with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
#     df_indicators.to_excel(writer, sheet_name="Indicators", index=False)
#     df_countries.to_excel(writer, sheet_name="Countries", index=False)
#     df_observations.to_excel(writer, sheet_name="Observations", index=False)
#
# print(f"模板已生成完毕：{output_file}")
import openpyxl
# import openpyxl
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import Font
#
# # Create a new workbook and select the active worksheet
# wb = openpyxl.Workbook()
# ws = wb.active
# ws.title = "GSD Calculations"
#
# # Define headers and input fields
# headers = [
#     ("Parameter", "Value", "Unit", "Formula Explanation"),
#     ("Flight Altitude (H)", "", "mm", "User input"),
#     ("Pixel Size (p)", "", "mm/px", "User input"),
#     ("Focal Length (f)", "", "mm", "User input"),
#     ("Resolution Width", "", "px", "User input"),
#     ("Resolution Height", "", "px", "User input"),
#     ("Forward Overlap (α)", "", "", "User input (e.g., 0.5 for 50%)"),
#     ("Side Overlap (β)", "", "", "User input (e.g., 0.2 for 20%)"),
#     ("Bit Depth", "", "bits", "User input"),
#     ("Number of Bands", "", "", "User input"),
# ]
#
# # Insert headers and input rows
# for row_idx, row_data in enumerate(headers, start=1):
#     for col_idx, cell_value in enumerate(row_data, start=1):
#         ws.cell(row=row_idx, column=col_idx, value=cell_value)
#
# # Define calculated values
# calculations = [
#     ("Ground Sample Distance (GSD)", "=B2*B3/B4", "m/px", "GSD = (H × p) / f"),
#     ("Footprint Width", "=B5*B12", "m", "Width = Resolution Width × GSD"),
#     ("Footprint Height", "=B6*B12", "m", "Height = Resolution Height × GSD"),
#     ("Coverage Area", "=B13*B14", "m²", "Area = Width × Height"),
#     ("Coverage Area in Hectares", "=B15/10000", "ha", "Hectares = Area / 10,000"),
#     ("Effective Coverage", "=B16*(1-B7)*(1-B8)", "ha", "A_effective = A_single × (1-α) × (1-β)"),
#     ("Images Required per Hectare", "=1/B17", "", "NumImages = 1 / A_effective"),
#     ("File Size per Image", "=(B5*B6*B9*B10)/(8*10^6)", "MB", "File Size = (Width × Height × Bit Depth × Bands) / (8 × 10^6)"),
#     ("Total Storage per Hectare", "=B18*B19", "MB", "Total Storage = NumImages × File Size"),
# ]
#
# # Insert calculated rows
# for row_idx, row_data in enumerate(calculations, start=len(headers) + 2):
#     for col_idx, cell_value in enumerate(row_data, start=1):
#         ws.cell(row=row_idx, column=col_idx, value=cell_value)
#
# # Apply bold font to headers
# bold_font = Font(bold=True)
# for col_idx in range(1, 5):
#     ws.cell(row=1, column=col_idx).font = bold_font
#
# # Auto-adjust column widths
# for col_idx in range(1, 5):
#     ws.column_dimensions[get_column_letter(col_idx)].width = 25
#
# # Save the file
# file_path = "GSD_Calculations_Template.xlsx"
# wb.save(file_path)
#
# # Provide the file path for download
# file_path
# Create an Excel file with formulas for automatic GSD, footprint, coverage, and storage calculations

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Create a new workbook and worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "GSD Calculations"

# Define headers and input fields for different sensors
headers = [
    ("Sensor Type", "Resolution Width (px)", "Resolution Height (px)", "GSD (m/px)", "Footprint Width (m)", "Footprint Height (m)",
     "Coverage Area (m²)", "Coverage Area (ha)", "Effective Coverage (ha)", "Images per Hectare", "File Size per Image (MB)", "Total Storage per Hectare (MB)"),
]

# Input data for different sensors
sensor_data = [
    ("Multispectral (MS)", 2064, 1544, 0.0528, "=B2*D2", "=C2*D2", "=E2*F2", "=G2/10000", "=H2*(1-0.5)*(1-0.2)", "=1/I2", "=(B2*C2*16*5)/(8*10^6)", "=J2*K2"),
    ("Panchromatic (PAN)", 4112, 3008, 0.0249, "=B3*D3", "=C3*D3", "=E3*F3", "=G3/10000", "=H3*(1-0.5)*(1-0.2)", "=1/I3", "=(B3*C3*16*1)/(8*10^6)", "=J3*K3"),
    ("Thermal (LWIR)", 320, 256, 0.335, "=B4*D4", "=C4*D4", "=E4*F4", "=G4/10000", "=H4*(1-0.5)*(1-0.2)", "=1/I4", "=(B4*C4*16*1)/(8*10^6)", "=J4*K4"),
]

# Insert headers
for row_idx, row_data in enumerate(headers, start=1):
    for col_idx, cell_value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=cell_value).font = Font(bold=True)

# Insert sensor data with formulas
for row_idx, row_data in enumerate(sensor_data, start=2):
    for col_idx, cell_value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=cell_value)

# Auto-adjust column widths
for col_idx in range(1, len(headers[0]) + 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = 22

# Save the Excel file
file_path = "GSD_Calculations_Automated.xlsx"
wb.save(file_path)

# Provide the file for download
file_path
